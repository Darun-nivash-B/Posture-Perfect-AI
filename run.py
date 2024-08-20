import logging
import math
import pandas as pd
import sys
import time
import openpyxl
from tf_pose import common
import cv2
import numpy as np
from tf_pose.estimator import TfPoseEstimator
from tf_pose.networks import get_graph_path, model_wh
import face_recognition
import os
import datetime
from datetime import datetime as dt
import matplotlib.pyplot as plt

logger = logging.getLogger('TfPoseEstimatorRun')
logger.handlers.clear()
logger.setLevel(logging.DEBUG)
ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)
formatter = logging.Formatter('[%(asctime)s] [%(name)s] [%(levelname)s] %(message)s')
ch.setFormatter(formatter)
logger.addHandler(ch)

def check_parameters(reference_parameters, new_parameters):
    for ref_val, new_val in zip(reference_parameters, new_parameters):
        if isinstance(ref_val, (int, float)) and isinstance(new_val, (int, float)):
            percentage_diff = abs(ref_val - new_val) / ref_val * 100
            if 1 < percentage_diff < 2:
                current_datetime = dt.now().strftime("%Y-%m-%d %H:%M:%S")
                message = f"Slightly Slouching - {current_datetime}"
                print(message)
            elif 2 <= percentage_diff < 3:
                current_datetime = dt.now().strftime("%Y-%m-%d %H:%M:%S")
                message = f"Moderately Slouching - {current_datetime}"
                print(message)
            elif percentage_diff >= 3:
                current_datetime = dt.now().strftime("%Y-%m-%d %H:%M:%S")
                message = f"Critically Slouching - {current_datetime}"
                print(message)
                break  # Exit the loop if a critical slouching message is shown
    else:
        print("No significant differences found.")

def read_excel_filter_columns(filename, names):
    df = pd.read_excel(filename)
    filtered_df = df[df['Name'].isin(names)]
    positions = filtered_df[['Position (x)', 'Position (y)']]
    return positions

def calculate_angle_between_points(point1, point2):
    vector = np.array(point2) - np.array(point1)
    angle_radians = np.arctan2(vector[1], vector[0])
    angle_degrees = np.degrees(angle_radians)
    return round(angle_degrees, 2)

def calculate_distance_between_points(point1, point2):
    point1 = np.array(point1)
    point2 = np.array(point2)
    distance = np.linalg.norm(point2 - point1)
    return round(distance, 4)
#Flask done
def take_picture():
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        logger.error("Error: Could not open camera.")
        return None
    ret, frame = cap.read()
    cap.release()
    return frame
#Flask done

def detect_faces(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    face_cascade = cv2.CascadeClassifier(r"C:\Users\reshm\OneDrive\Desktop\face_recogniion\haarcascade_frontalface_default.xml")
    faces = face_cascade.detectMultiScale(gray, 1.3, 6)
    print('Number of detected faces:', len(faces))
    if len(faces) > 0:
        for i, (x, y, w, h) in enumerate(faces):
            cv2.rectangle(image, (x, y), (x + w, y + h), (0, 255, 255), 2)
            face = image[y:y + h, x:x + w]
            cv2.imwrite(f'C:/Project/myWorkspace/tf-pose-estimation/images/detectedFaces/{i}.jpg', face)
            print(f"{i}.jpg is saved")
    return faces

def match_faces_with_database(unknown_image_path, know_image_path):
    matched_image = None
    for item in unknown_image_path:
        for pic in know_image_path:
            known_image = face_recognition.load_image_file('C:/Project/myWorkspace/tf-pose-estimation/images/Database Faces/'+pic)
            known_encoding = face_recognition.face_encodings(known_image)[0]
            unknown_image = face_recognition.load_image_file('C:/Project/myWorkspace/tf-pose-estimation/images/detectedFaces/'+item)
            unknown_encoding = face_recognition.face_encodings(unknown_image)[0]
            results = face_recognition.api.compare_faces([known_encoding], unknown_encoding, tolerance=0.5)
            logger.debug(results)
            if results[0]:
                matched_image = pic.split('.')[0]
                break
        if matched_image:
            break
    return matched_image

def analyze_posture(path, resize='432x368', model='mobilenet_thin'):
    w, h = model_wh(resize)
    if w == 0 or h == 0:
        e = TfPoseEstimator(get_graph_path(model), target_size=(432, 368))
    else:
        e = TfPoseEstimator(get_graph_path(model), target_size=(w, h))

    image = common.read_imgfile(path, None, None)
    if image is None:
        logger.error('Image can not be read, path=%s' % path)
        return

    t = datetime.datetime.now()
    humans = e.inference(image, resize_to_default=(w > 0 and h > 0), upsample_size=4)
    elapsed = datetime.datetime.now() - t
    logger.debug('inference image: %s in %.4f seconds.' % (path, elapsed.total_seconds()))

    return humans, image

def write_posture_details_to_excel(matched_image, humans, workbook_path):
    current_time_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook.active
    if worksheet.max_row == 1:
        worksheet.append(["Human", "Body Part", "Name", "Score", "Position (x)", "Position (y)", "Time Stamp"])
    if len(humans) == 0:
        print("No humans detected.")
    else:
        for i, human in enumerate(humans):
            for j, body_part in human.body_parts.items():
                body_part_name = common.CocoPart(body_part.part_idx).name
                score = round(body_part.score, 2)
                position_x = round(body_part.x, 2)
                position_y = round(body_part.y, 2)
                worksheet.append([matched_image, j, body_part_name, score, position_x, position_y, current_time_str])
    #only run for the list of names we choose and break when the  list is done
    workbook.save(workbook_path)

def clear_excel_sheet(file_path, sheet_name):
    try:
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' does not exist in the workbook.")
            return
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None
        workbook.save(file_path)
        print("Excel sheet cleared successfully.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

def plot_poses(image, humans):
    image = TfPoseEstimator.draw_humans(image, humans, imgcopy=False)

    fig = plt.figure()

    a = fig.add_subplot(2, 2, 1)
    a.set_title('Detected Poses')
    plt.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))

    try:
        bgimg = cv2.cvtColor(image.astype(np.uint8), cv2.COLOR_BGR2RGB)
        bgimg = cv2.resize(bgimg, (e.heatMat.shape[1], e.heatMat.shape[0]), interpolation=cv2.INTER_AREA)
        a = fig.add_subplot(2, 2, 2)
        plt.imshow(bgimg, alpha=0.5)
        tmp = np.amax(e.heatMat[:, :, :-1], axis=2)
        plt.imshow(tmp, cmap=plt.cm.gray, alpha=0.5)
        plt.colorbar()
        a.set_title('Heatmap')

        tmp2 = e.pafMat.transpose((2, 0, 1))
        tmp2_odd = np.amax(np.absolute(tmp2[::2, :, :]), axis=0)
        a = fig.add_subplot(2, 2, 3)
        plt.imshow(tmp2_odd, cmap=plt.cm.gray, alpha=0.5)
        plt.colorbar()
        a.set_title('Vector Map - X')

        tmp2_even = np.amax(np.absolute(tmp2[1::2, :, :]), axis=0)
        a = fig.add_subplot(2, 2, 4)
        plt.imshow(tmp2_even, cmap=plt.cm.gray, alpha=0.5)
        plt.colorbar()
        a.set_title('Vector Map - Y')

        plt.tight_layout()
        plt.show()
    except Exception as e:
        logger.warning('matplotlib error, %s' % e)
        cv2.imshow('result', image)
        cv2.waitKey()

if __name__ == '__main__':
    unknown_image_path = os.listdir(r'C:\Project\myWorkspace\tf-pose-estimation\images\DetectedFaces')
    know_image_path = os.listdir(r"C:\Project\myWorkspace\tf-pose-estimation\images\Database Faces")
    resize = '432x368'
    model = 'mobilenet_thin'
    n = 1
    while True:
        if n == 1:
            picture = take_picture()
            save_directory = r"C:\Users\reshm\OneDrive\Desktop\Project work"
            filename = os.path.join(save_directory, "reference.jpg")
            cv2.imwrite(filename, picture)
            print(f"Picture taken and saved as {filename}")

            path = r"C:\Project\myWorkspace\tf-pose-estimation\images\darunStraight.jpg"
            image = cv2.imread(path)
            humans, image = analyze_posture(path, resize, model)
            faces = detect_faces(image)
            logger.debug('Number of detected faces: %d' % len(faces))
            matched_image = match_faces_with_database(unknown_image_path, know_image_path)
            if matched_image:
                logger.debug("Match found! The matched image is: %s" % matched_image)
            else:
                logger.debug("No match found.")
            write_posture_details_to_excel(matched_image, humans, r"C:\Users\reshm\OneDrive\Desktop\PostureDetails.xlsx")
            logger.debug("Updated Excel.")

            filename = r"C:\Users\reshm\OneDrive\Desktop\PostureDetails.xlsx"
            names_to_filter = ['Neck', 'RHip', 'LHip', 'Nose', 'RShoulder', 'LShoulder']
            positions = read_excel_filter_columns(filename, names_to_filter)
            nose = [positions.iloc[0]['Position (x)'], positions.iloc[0]['Position (y)']]
            neck = [positions.iloc[1]['Position (x)'], positions.iloc[1]['Position (y)']]
            rShoulder = [positions.iloc[2]['Position (x)'], positions.iloc[2]['Position (y)']]
            lShoulder = [positions.iloc[3]['Position (x)'], positions.iloc[3]['Position (y)']]
            rHip = [positions.iloc[4]['Position (x)'], positions.iloc[4]['Position (y)']]
            lHip = [positions.iloc[5]['Position (x)'], positions.iloc[5]['Position (y)']]
            angleP1 = calculate_angle_between_points(nose, rShoulder)
            angleP2 = calculate_angle_between_points(nose, lShoulder)
            angleP3 = calculate_angle_between_points(neck, rHip)
            angleP4 = calculate_angle_between_points(neck, lHip)
            print(angleP1, angleP2, angleP3, angleP4)
            distanceP1 = calculate_distance_between_points(nose, rShoulder)
            distanceP2 = calculate_distance_between_points(nose, lShoulder)
            distanceP3 = calculate_distance_between_points(neck, rHip)
            distanceP4 = calculate_distance_between_points(neck, lHip)
            print(distanceP1, distanceP2, distanceP3, distanceP4)
            time.sleep(5)
            n = n + 1
        clear_excel_sheet("C:/Users/reshm/OneDrive/Desktop/PostureDetails.xlsx", 'Sheet1')
        if n == 2:
            while True:
                picture = take_picture()
                save_directory = r"C:\Users\reshm\OneDrive\Desktop\Project work"
                filename = os.path.join(save_directory, "test.jpg")
                cv2.imwrite(filename, picture)
                print(f"Picture taken and saved as {filename}")
                path = r"C:\Project\myWorkspace\tf-pose-estimation\images\Anish_Posture.jpg"
                humans, image = analyze_posture(path, resize, model)
                pic = cv2.imread(path)
                faces = detect_faces(pic)
                logger.debug('Number of detected faces: %d' % len(faces))
                matched_image = match_faces_with_database(unknown_image_path, know_image_path)
                if matched_image:
                    logger.debug("Match found! The matched image is: %s" % matched_image)
                else:
                    logger.debug("No match found.")
                write_posture_details_to_excel(matched_image, humans, r"C:\Users\reshm\OneDrive\Desktop\PostureDetails.xlsx")
                logger.debug("Updated Excel.")

                filename = r"C:\Users\reshm\OneDrive\Desktop\PostureDetails.xlsx"
                names_to_filter = ['Neck', 'RHip', 'LHip', 'Nose', 'RShoulder', 'LShoulder']
                positions = read_excel_filter_columns(filename, names_to_filter)
                noseP = [positions.iloc[0]['Position (x)'], positions.iloc[0]['Position (y)']]
                neckP = [positions.iloc[1]['Position (x)'], positions.iloc[1]['Position (y)']]
                rShoulderP = [positions.iloc[2]['Position (x)'], positions.iloc[2]['Position (y)']]
                lShoulderP = [positions.iloc[3]['Position (x)'], positions.iloc[3]['Position (y)']]
                rHipP = [positions.iloc[4]['Position (x)'], positions.iloc[4]['Position (y)']]
                lHipP = [positions.iloc[5]['Position (x)'], positions.iloc[5]['Position (y)']]
                angleP1P = calculate_angle_between_points(noseP, rShoulderP)
                angleP2P = calculate_angle_between_points(noseP, lShoulderP)
                angleP3P = calculate_angle_between_points(neckP, rHipP)
                angleP4P = calculate_angle_between_points(neckP, lHipP)
                distanceP1P = calculate_distance_between_points(noseP, rShoulderP)
                distanceP2P = calculate_distance_between_points(noseP, lShoulderP)
                distanceP3P = calculate_distance_between_points(neckP, rHipP)
                distanceP4P = calculate_distance_between_points(neckP, lHipP)
                clear_excel_sheet("C:/Users/reshm/OneDrive/Desktop/PostureDetails.xlsx", 'Sheet1')
                check_parameters([distanceP1, distanceP2, distanceP3, distanceP4], [distanceP1P, distanceP2P, distanceP3P, distanceP4P])
                check_parameters([angleP1, angleP2, angleP3, angleP4], [angleP1P, angleP2P, angleP3P, angleP4P])
                time.sleep(10)
